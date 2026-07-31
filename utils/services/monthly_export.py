import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


class AttendanceExcelExportService:
    @staticmethod
    def format_money(value):
        try:
            return f"{int(value):,}".replace(",", " ")
        except Exception:
            return value

    @staticmethod
    def auto_adjust_column_width(ws):
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column
            column_letter = get_column_letter(column)

            for cell in column_cells:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 3

    @staticmethod
    def style_sheet(ws):
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        for cell in ws[1]:
            cell.font = Font(bold=True)

    @staticmethod
    def generate_monthly_excel(report, year, month):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Oylik hisobot"
        ws.append(
            [
                "ID",
                "Hodimlar",
                "Sababli kelmadi",
                "Sababsiz kelmadi",
                "Ishlagan soati",
                "Smena",
                "Maosh",
                "Bonus",
                "Jarima",
                "Hisoblangan miqdor",
                "Yangi maosh",
            ]
        )

        total_salary = 0
        total_bonus = 0
        total_penalty = 0

        for row in report["results"]:
            total_salary += row["employee_salary"]
            total_bonus += row["total_bonus"]
            total_penalty += row["total_penalty"]

            ws.append(
                [
                    row["employee_id"],
                    row["employee_name"],
                    row["sbk_count"],
                    row["szk_count"],
                    row["worked_time"],
                    f"{row['shift_start_time']} - {row['shift_end_time']}",
                    AttendanceExcelExportService.format_money(row["employee_salary"]),
                    AttendanceExcelExportService.format_money(row["total_bonus"]),
                    AttendanceExcelExportService.format_money(row["total_penalty"]),
                    AttendanceExcelExportService.format_money(row["net_adjustment"]),
                    AttendanceExcelExportService.format_money(row["new_salary"]),
                ]
            )

        ws.append([])
        ws.append(["Jami hodimlar", report["count"]])
        ws.append(["Umumiy maosh", AttendanceExcelExportService.format_money(total_salary)])
        ws.append(["Jami bonus", AttendanceExcelExportService.format_money(total_bonus)])
        ws.append(["Jami jarima", AttendanceExcelExportService.format_money(total_penalty)])
        AttendanceExcelExportService.style_sheet(ws)
        AttendanceExcelExportService.auto_adjust_column_width(ws)
        details_ws = wb.create_sheet("Daily Details")
        details_ws.append(
            [
                "Hodimlar",
                "Sana",
                "Status",
                "Birinchi kirish",
                "Oxirgi chiqish",
                "Ishladi",
                "Farq",
                "Jarima",
                "Bonus",
                "Kunlik Jami",
            ]
        )

        for emp in report["results"]:
            for detail in emp["details"]:
                details_ws.append(
                    [
                        emp["employee_name"],
                        detail.get("date"),
                        detail.get("status_label"),
                        detail.get("first_in"),
                        detail.get("last_out"),
                        detail.get("worked"),
                        detail.get("difference"),
                        AttendanceExcelExportService.format_money(detail.get("penalty")),
                        AttendanceExcelExportService.format_money(detail.get("bonus")),
                        AttendanceExcelExportService.format_money(detail.get("daily_total")),
                    ]
                )

        AttendanceExcelExportService.style_sheet(details_ws)
        AttendanceExcelExportService.auto_adjust_column_width(details_ws)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename=attendance_{year}_{month}.xlsx"

        wb.save(response)
        return response
