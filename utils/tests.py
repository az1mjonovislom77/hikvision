from datetime import date, datetime, time, timedelta
from io import BytesIO, StringIO

import openpyxl
from django.core.cache import cache
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.utils import timezone
from django.utils.timezone import make_aware
from rest_framework.test import APIClient

from attendance.models import AttendanceDaily
from attendance.tasks.mark_attendance import mark_daily_attendance
from day.models import Shift
from event.models import AccessEvent
from person.models import Employee
from user.models import User
from utils.models import Branch, Devices, Notification, Plan, Subscription
from utils.services.monthly_export import AttendanceExcelExportService
from utils.services.notifications import NotificationService
from utils.services.smartcity_daily_stats import SmartCityDailyStatsService
from utils.services.smartcity_stats import SmartCityStatsService
from utils.services.subscription import SubscriptionService
from utils.utils.rate_limit import check_login_rate_limit, reset_login_rate_limit


class MarkAttendanceTaskTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(phone_number="998904000001", password="pass")
        self.device = Devices.objects.create(
            user=self.user,
            name="D",
            ip="10.0.0.3",
            username="admin",
            password="admin",
            status=Devices.Status.ACTIVE,
        )
        self.emp_present = Employee.objects.create(device=self.device, employee_no="P1", name="Present Employee")
        self.emp_absent = Employee.objects.create(device=self.device, employee_no="A1", name="Absent Employee")

    def test_present_employee_marked_present(self):
        target = date(2026, 6, 10)
        AccessEvent.objects.create(
            employee=self.emp_present,
            serial_no="s1",
            time="2026-06-10T09:00:00+05:00",
            major=5,
            minor=75,
            major_name="m",
            minor_name="n",
            name="P1",
            employee_no="P1",
            picture_url="",
            raw_json={},
            device=self.device,
            label_name="KIRISH",
        )
        mark_daily_attendance(target_date_str=str(target))
        rec = AttendanceDaily.objects.get(employee=self.emp_present, date=target)
        self.assertEqual(rec.status, "present")

    def test_absent_employee_marked_absent(self):
        target = date(2026, 6, 10)
        mark_daily_attendance(target_date_str=str(target))
        rec = AttendanceDaily.objects.get(employee=self.emp_absent, date=target)
        self.assertEqual(rec.status, "absent_unexcused")

    def test_task_is_idempotent(self):
        target = date(2026, 6, 11)
        mark_daily_attendance(target_date_str=str(target))
        mark_daily_attendance(target_date_str=str(target))
        count = AttendanceDaily.objects.filter(employee=self.emp_absent, date=target).count()
        self.assertEqual(count, 1)


@override_settings(SECURE_SSL_REDIRECT=False)
class AttendanceAPITests(TestCase):
    client: APIClient

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(phone_number="998904000002", password="pass")
        self.device = Devices.objects.create(
            user=self.user,
            name="D2",
            ip="10.0.0.4",
            username="admin",
            password="admin",
            status=Devices.Status.ACTIVE,
        )
        self.branch = Branch.objects.create(user=self.user, name="B1", device=self.device)
        self.client.force_authenticate(user=self.user)

    def test_absent_view_requires_branch_id(self):
        response = self.client.get("/attendance/absent/")
        self.assertEqual(response.status_code, 400)

    def test_absent_view_invalid_date_returns_400(self):
        response = self.client.get(f"/attendance/absent/?branch_id={self.branch.id}&date=not-a-date")
        self.assertEqual(response.status_code, 400)

    def test_monthly_report_invalid_year_returns_400(self):
        response = self.client.get(f"/attendance/report/monthly/?branch_id={self.branch.id}&year=abc&month=1")
        self.assertEqual(response.status_code, 400)


class PlanModelTests(TestCase):
    def test_duration_months_derived_from_cycle(self):
        plan = Plan.objects.create(
            title="Go", plan_type=Plan.PlanType.GO, billing_cycle=Plan.CycleChoice.QUARTERLY, price=10
        )
        self.assertEqual(plan.duration_months, 3)

    def test_yearly_cycle_is_twelve_months(self):
        plan = Plan.objects.create(
            title="Plus", plan_type=Plan.PlanType.PLUS, billing_cycle=Plan.CycleChoice.YEARLY, price=100
        )
        self.assertEqual(plan.duration_months, 12)
        self.assertEqual(str(plan), "Plus (12 months)")


class SubscriptionModelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(phone_number="998909000001", password="pass")
        self.plan = Plan.objects.create(
            title="Go", plan_type=Plan.PlanType.GO, billing_cycle=Plan.CycleChoice.MONTHLY, price=10
        )

    def test_remaining_days_future(self):
        sub = Subscription.objects.create(
            user=self.user,
            plan=self.plan,
            start_date=timezone.now(),
            end_date=timezone.now() + timedelta(days=10, hours=1),
            is_active=True,
        )
        self.assertEqual(sub.remaining_days, 10)

    def test_remaining_days_past_is_zero(self):
        sub = Subscription.objects.create(
            user=self.user,
            plan=self.plan,
            start_date=timezone.now() - timedelta(days=40),
            end_date=timezone.now() - timedelta(days=3),
            is_active=False,
        )
        self.assertEqual(sub.remaining_days, 0)


class SubscriptionServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(phone_number="998909000002", password="pass")
        self.plan = Plan.objects.create(
            title="Go", plan_type=Plan.PlanType.GO, billing_cycle=Plan.CycleChoice.MONTHLY, price=10
        )

    def test_deactivate_previous(self):
        sub = Subscription.objects.create(
            user=self.user,
            plan=self.plan,
            start_date=timezone.now(),
            end_date=timezone.now() + timedelta(days=30),
            is_active=True,
        )
        SubscriptionService.deactivate_previous(self.user)
        sub.refresh_from_db()
        self.assertFalse(sub.is_active)

    def test_resolve_target_user_regular_user_returns_self(self):
        other = User.objects.create_user(phone_number="998909000003", password="pass")
        self.assertEqual(SubscriptionService.resolve_target_user(self.user, other.id), self.user)

    def test_resolve_target_user_staff_fetches_by_id(self):
        staff = User.objects.create_user(phone_number="998909000004", password="pass", is_staff=True)
        self.assertEqual(SubscriptionService.resolve_target_user(staff, self.user.id), self.user)


class NotificationServiceTests(TestCase):
    def setUp(self):
        self.user1 = User.objects.create_user(phone_number="998909000005", password="pass")
        self.user2 = User.objects.create_user(phone_number="998909000006", password="pass")

    def test_send_bulk_creates_notifications(self):
        count = NotificationService.send_bulk("Salom", [self.user1, self.user2])
        self.assertEqual(count, 2)
        self.assertEqual(Notification.objects.filter(text="Salom").count(), 2)

    def test_resolve_users_filters_by_ids(self):
        users = NotificationService.resolve_users([self.user1.id])
        self.assertEqual(list(users), [self.user1])


class NotifySubscriptionsCommandTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(phone_number="998909000007", password="pass")
        self.plan = Plan.objects.create(
            title="Go", plan_type=Plan.PlanType.GO, billing_cycle=Plan.CycleChoice.MONTHLY, price=10
        )

    def test_creates_notifications_for_subscriptions_expiring_in_a_week(self):
        Subscription.objects.create(
            user=self.user,
            plan=self.plan,
            start_date=timezone.now() - timedelta(days=23),
            end_date=timezone.now() + timedelta(days=7),
            is_active=True,
        )
        Subscription.objects.create(
            user=self.user,
            plan=self.plan,
            start_date=timezone.now() - timedelta(days=29),
            end_date=timezone.now() + timedelta(days=1),
            is_active=True,
        )

        out = StringIO()
        call_command("notify_subscriptions", stdout=out)

        self.assertEqual(Notification.objects.filter(user=self.user).count(), 1)
        self.assertIn("1 ta notification", out.getvalue())


class LoginRateLimitTests(TestCase):
    def setUp(self):
        cache.clear()

    def test_limit_blocks_after_seven_attempts(self):
        for _ in range(7):
            self.assertTrue(check_login_rate_limit("1.2.3.4", "998900000000"))
        self.assertFalse(check_login_rate_limit("1.2.3.4", "998900000000"))

    def test_reset_allows_again(self):
        for _ in range(8):
            check_login_rate_limit("1.2.3.4", "998900000000")
        reset_login_rate_limit("1.2.3.4", "998900000000")
        self.assertTrue(check_login_rate_limit("1.2.3.4", "998900000000"))


class SmartCityStatsServiceTests(TestCase):
    def test_invalid_time_filter_raises(self):
        with self.assertRaises(ValueError):
            SmartCityStatsService("month")

    def test_build_counts_present_and_absent(self):
        mahalla = User.objects.create_user(phone_number="998909000008", password="pass", role=User.UserRoles.MAHALLA)
        device = Devices.objects.create(
            user=mahalla, name="MD", ip="10.0.0.20", username="a", password="a", status=Devices.Status.ACTIVE
        )
        Employee.objects.create(
            device=device, employee_no="S1", name="Present", begin_time=timezone.now() - timedelta(days=1)
        )
        Employee.objects.create(device=device, employee_no="S2", name="Absent")

        data = SmartCityStatsService("week").build()

        self.assertEqual(data["summary"]["totalEmployees"], 2)
        self.assertEqual(data["summary"]["presentEmployees"], 1)
        self.assertEqual(data["summary"]["absentEmployees"], 1)
        self.assertEqual(data["summary"]["averageAttendance"], 50.0)


class SmartCityDailyStatsServiceTests(TestCase):
    def test_invalid_date_raises(self):
        with self.assertRaises(ValueError):
            SmartCityDailyStatsService("bad-date")

    def test_detail_statuses(self):
        mahalla = User.objects.create_user(phone_number="998909000009", password="pass", role=User.UserRoles.MAHALLA)
        device = Devices.objects.create(
            user=mahalla, name="MD2", ip="10.0.0.21", username="a", password="a", status=Devices.Status.ACTIVE
        )
        shift = Shift.objects.create(user=mahalla, name="M-shift", start_time=time(9, 0), end_time=time(18, 0))
        late = Employee.objects.create(
            device=device,
            employee_no="L1",
            name="Late",
            shift=shift,
            begin_time=make_aware(datetime(2026, 6, 10, 15, 0)),
        )
        on_time = Employee.objects.create(
            device=device,
            employee_no="O1",
            name="OnTime",
            shift=shift,
            begin_time=make_aware(datetime(2026, 6, 10, 8, 30)),
        )
        absent = Employee.objects.create(device=device, employee_no="A1", name="Absent", shift=shift)

        data = SmartCityDailyStatsService("2026-06-10", mahalla_id=mahalla.id).build()

        self.assertEqual(data["type"], "detail")
        statuses = {row["employeeId"]: row["status"] for row in data["data"]}
        self.assertEqual(statuses[late.id], "LATE")
        self.assertEqual(statuses[on_time.id], "ON_TIME")
        self.assertEqual(statuses[absent.id], "ABSENT")


class MonthlyExcelExportTests(TestCase):
    def test_generate_monthly_excel_returns_xlsx(self):
        report = {
            "count": 1,
            "results": [
                {
                    "employee_id": 1,
                    "employee_name": "A",
                    "sbk_count": 0,
                    "szk_count": 1,
                    "worked_time": "8:00",
                    "shift_start_time": "09:00",
                    "shift_end_time": "18:00",
                    "employee_salary": 100,
                    "total_bonus": 5,
                    "total_penalty": 3,
                    "net_adjustment": 2,
                    "new_salary": 102,
                    "details": [
                        {
                            "date": "2026-06-10",
                            "status_label": "Ishlagan",
                            "first_in": "09:00",
                            "last_out": "18:00",
                            "worked": "8:00",
                            "difference": "0:00",
                            "penalty": 0,
                            "bonus": 0,
                            "daily_total": 0,
                        }
                    ],
                }
            ],
        }

        response = AttendanceExcelExportService.generate_monthly_excel(report, 2026, 6)

        self.assertEqual(response["Content-Disposition"], "attachment; filename=attendance_2026_6.xlsx")
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        self.assertIn("Oylik hisobot", workbook.sheetnames)
        self.assertIn("Daily Details", workbook.sheetnames)
        sheet = workbook["Oylik hisobot"]
        self.assertEqual(sheet.cell(row=2, column=2).value, "A")

    def test_format_money(self):
        self.assertEqual(AttendanceExcelExportService.format_money(1234567), "1 234 567")
        self.assertEqual(AttendanceExcelExportService.format_money("n/a"), "n/a")


@override_settings(SECURE_SSL_REDIRECT=False)
class DevicePasswordExposureTests(TestCase):
    client: APIClient

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(phone_number="998907300001", password="pass")
        self.device = Devices.objects.create(
            user=self.user,
            name="D",
            ip="10.0.0.30",
            username="admin",
            password="secret123",
            status=Devices.Status.ACTIVE,
        )

    def test_device_password_is_not_in_response(self):
        self.client.force_authenticate(self.user)

        response = self.client.get("/utils/devices/")

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("password", response.data[0])

    def test_device_password_is_still_writable(self):
        self.client.force_authenticate(self.user)

        response = self.client.put(f"/utils/devices/{self.device.id}/", {"password": "newsecret"}, format="json")

        self.assertEqual(response.status_code, 200)
        self.device.refresh_from_db()
        self.assertEqual(self.device.password, "newsecret")


@override_settings(SECURE_SSL_REDIRECT=False)
class AdminNotificationPermissionTests(TestCase):
    client: APIClient

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(phone_number="998907400001", password="pass")
        self.other = User.objects.create_user(phone_number="998907400002", password="pass")
        self.admin = User.objects.create_user(
            phone_number="998907400003", password="pass", role=User.UserRoles.SUPERADMIN
        )

    def test_regular_user_cannot_broadcast(self):
        self.client.force_authenticate(self.user)

        response = self.client.post("/utils/admin/notification/", {"text": "spam"}, format="json")

        self.assertEqual(response.status_code, 403)
        self.assertEqual(Notification.objects.count(), 0)

    def test_superadmin_can_broadcast(self):
        self.client.force_authenticate(self.admin)

        response = self.client.post("/utils/admin/notification/", {"text": "elon"}, format="json")

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Notification.objects.filter(text="elon").count(), 3)


@override_settings(SECURE_SSL_REDIRECT=False)
class PlanPermissionTests(TestCase):
    client: APIClient

    def setUp(self):
        self.client = APIClient()
        self.plan = Plan.objects.create(
            title="Go", plan_type=Plan.PlanType.GO, billing_cycle=Plan.CycleChoice.MONTHLY, price=10
        )
        self.user = User.objects.create_user(phone_number="998907500001", password="pass")
        self.admin = User.objects.create_user(
            phone_number="998907500002", password="pass", role=User.UserRoles.SUPERADMIN
        )

    def test_anonymous_can_read_plans(self):
        response = self.client.get("/utils/plan/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)

    def test_anonymous_cannot_delete_plan(self):
        response = self.client.delete(f"/utils/plan/{self.plan.id}/")

        self.assertIn(response.status_code, (401, 403))
        self.assertTrue(Plan.objects.filter(id=self.plan.id).exists())

    def test_regular_user_cannot_create_plan(self):
        self.client.force_authenticate(self.user)

        response = self.client.post(
            "/utils/plan/", {"title": "Hack", "plan_type": "go", "billing_cycle": "monthly", "price": 1}, format="json"
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(Plan.objects.count(), 1)

    def test_admin_can_create_plan(self):
        self.client.force_authenticate(self.admin)

        response = self.client.post(
            "/utils/plan/", {"title": "New", "plan_type": "plus", "billing_cycle": "yearly", "price": 5}, format="json"
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Plan.objects.count(), 2)
